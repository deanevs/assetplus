import openpyxl
from pathlib import Path


def main():

    file_path = Path(r"C:\Users\212628255\Documents\2 GE\AssetPlus\3 Sites\2 Spire\Nuvolo")
    file_input = "Nuvolo upload - week 1-6 NG edit .xlsx"
    output = Path(r"C:\Users\212628255\Documents\2 GE\AssetPlus\3 Sites\2 Spire\Nuvolo\output")

    sheet_to_read = 'Assets '

    wb = openpyxl.load_workbook(file_path / file_input)
    print(wb.sheetnames)
    sheet = wb[sheet_to_read]

    catch = True
    line_count = 0
    file_count = 0



    for row in sheet.iter_rows(values_only=True):
        if catch:       # ignore headers
            catch = False
            continue
        asset_id = row[1]
        asset_tag = row[2]
        if asset_tag is not None:
            print(f"Asset id = {asset_id} and asset_tag = {asset_tag}")
            line_count += 1
            if line_count == 1:     # first one
                file_count += 1
                sql_file = "Update_ " + str(file_count) + '.sql'
                fd = open(output / sql_file, 'w')
            elif line_count % 1000 == 0:
                fd.close()
                file_count += 1
                sql_file = "Update_" + str(file_count) + '.sql'
                fd = open(output / sql_file, 'w')
            fd.write(set_internal_asset_no(asset_id, asset_tag))

    fd.close()
    print(f"Added {line_count} rows")


def set_internal_asset_no(asset, tag):
    sql = f"-- {asset}\n" \
          f"UPDATE B_EQ1996 SET NUMBER_IN_SITE = '{tag}' WHERE N_IMMA = '{asset}'\n"
    return sql


if __name__ == '__main__':
    main()