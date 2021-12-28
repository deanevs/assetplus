import openpyxl
from pathlib import Path
import datetime


"""
D_DER_I_P = '2021-01-13'    # date of last PM
D_PRO_I_P = '20200318'      # date of next PM
"""

def main():

    # fn = Path(r"C:\Users\212628255\Documents\2 GE\AssetPlus\7 Projects\20210831 - LCS PM Dates\LCS AUGUST PM DATE ADJUSTMENTS.xlsx")
    # fn = Path(r"C:\Users\212628255\Documents\2 GE\AssetPlus\7 Projects\20210831 - LCS PM Dates\DI ULS LAST PM END SEP 21.xlsx")
    fn = Path(r"C:\Users\212628255\Documents\2 GE\AssetPlus\7 Projects\20210831 - LCS PM Dates\excludes.xlsx")

    wb = openpyxl.load_workbook(fn)
    print(wb.sheetnames)
    sheet = wb['excludes']

    catch = True
    for row in sheet.iter_rows(values_only=True):
        if catch:
            catch = False
            continue
        asset_id = row[0][1:]
        last_pm = date_to_str_last(row[8])
        next_pm = date_to_str_next(row[11])
        print(update_last_pm_eq1996(asset_id, last_pm, next_pm))


def update_last_pm_eq1996(asset, last, next):

    sql = f"-- {asset}\n" \
          f"UPDATE B_EQ1996\n" \
          f"SET D_DER_I_P = '{last}', D_PRO_I_P = '{next}' WHERE N_IMMA = '{asset}'\n"

    return sql


def date_to_str_last(dt):
    return dt.strftime("%Y-%m-%d")

def date_to_str_next(dt):
    return dt.strftime("%Y%m%d")

def clean_date(dt):
    y = dt[:3]
    m = dt[5:7]
    d = dt[8:10]
    return y + '-'+ m + '-' + d


if __name__ == '__main__':
    main()