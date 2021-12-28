from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, Alignment, Border, Side


def analyse(df):
    wb = Workbook()
    ws = wb.active

    # fill sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_map = get_header_map(ws)
    apply_header_format(ws)

    # conditional formatting
    # if % > 89 then highlight red the match, count and percent column
    red_background = PatternFill(bgColor=colors.RED)
    diff_style = DifferentialStyle(fill=red_background)
    rule1 = Rule(type="expression", dxf=diff_style)
    rule1.formula = ["$E2>89"]
    # set range
    number_rows = len(ws['E'])
    range = 'E2:E' + str(number_rows)
    # add formatting
    ws.conditional_formatting.add(range, rule1)

    format(ws)

    wb.save("C:\\Users\\212628255\\Documents\\GE\\AssetPlus\\7 Projects\\GMDN\\" + 'Analysis1.xlsx')

def format(ws):
    ws.title = 'Full Analysis'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 35
    ws.column_dimensions['J'].width = 6
    ws.column_dimensions['K'].width = 6

def get_header_map(ws):
    header_map = {}
    for row in ws.iter_rows(1, 1):
        for i, ref in enumerate(row):
            header_map[ref.value] = i
    return header_map


def apply_header_format(sheet):
    header_row = sheet[1]
    for cell in header_row:
       cell.font = Font(bold=True)
       cell.border = Border(bottom=Side(border_style="thin"))
       cell.alignment = Alignment(horizontal="center", vertical="center")










